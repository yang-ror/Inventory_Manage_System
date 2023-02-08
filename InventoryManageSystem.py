from openpyxl import Workbook, load_workbook
from colorama import Fore, Back, Style
import datetime
import logging


def main():

    logging.basicConfig(
		level = logging.DEBUG,
		format = "{asctime} {levelname:<8} {message}",
		datefmt='%Y-%m-%d %H:%M:%S',
		style = '{',
		filename = 'inventory.log',
		filemode = 'w'
	)

    testInTerminal()

def testInTerminal():
    # check file, unless all entries are valid, the program will not contine
    errorList = validateFile()
    print(f'Checking file, found {len(errorList)} error(s)')
    while len(errorList) > 0:
        print(Fore.WHITE + 'Errors Found:')
        print()
        for error in errorList:
            print(Fore.RED + f"Type: {error['type']}" + Style.RESET_ALL)
            print(Fore.MAGENTA + f"Location: {error['loc']}" + Style.RESET_ALL)
            print(Fore.YELLOW + f"SKU: {error['SKU']}" + Style.RESET_ALL)
            print()
        return
        
        x = input('Enter 1 to check again, 0 to exit')

        if x == '1':
            errorList = validateFile()

    # endless loop 
    while True:
        #take user input
        print(Fore.CYAN + 'Enter 1 to generate inventory report ' + Fore.YELLOW + ' 2 to generate shipment report ' + Fore.RED + ' 0 to exit')
        sku = input(Fore.GREEN + 'Enter SKU: ')

        #logs user SKU input
        # message = sku
        # logging.info(sku)

        while sku == '':
            sku = input(Fore.GREEN + 'Enter SKU: ')
        
        message = f"User Input SKU: {sku}"
        logging.info(message)

        if sku == '0':
            #end the loop
            break

        if sku == '1':
            inventoryReportGenerator()
            #go to the next loop
            continue

        # if sku == '2':
        #     shipmentReportGenerator()
        #     continue

        # find the formal SKU
        sku = parseSKU(sku)

        locationList = searchInventory(sku)

        if len(locationList) == 0:
            print(Fore.RED + 'found 0 items')
            print(Style.RESET_ALL)
            continue

        quantity = input(Fore.CYAN + 'Enter quantity (1): ')

        if quantity == '':
            quantity = 1

        if quantity == '0':
            print(Style.RESET_ALL)
            continue
        
        #logs user input quantity
        message = f"User Input Quantity: {quantity}"
        logging.info(message)

        total = 0
        for i in range(0, len(locationList)):
            color = Fore.MAGENTA
            if i%2 == 0: color = Fore.YELLOW
            print(color + f'({i+1}) ', locationList[i]['loc'], ':', locationList[i]['qty'], 'units' + Style.RESET_ALL)
            total += locationList[i]['qty']
        
        print('Total: ', total, 'units')

        while True:
            selected_loc = input(Fore.CYAN + 'Select a location (1): ')

            if selected_loc == '':
                selected_loc = '1'

            #logs user input location
            message = f"User Input Location: {selected_loc}"
            logging.info(message)
            
            try:
                locationList[int(selected_loc) - 1]

            except:
                message = 'Invalid Location'
                logging.warning(message)
                print(Back.RED + Fore.WHITE + Style.BRIGHT + '✗ INVALID LOCATION ' + Style.RESET_ALL)
            else:
                break
        
        if selected_loc == '0':
            print(Style.RESET_ALL)
            continue


        verified = False
        upc = None
        # endless loop until the user scanned the correct UPC
        while not verified and upc != '0':
            upc = input(Fore.BLUE + 'Scan the barcode: ')

            #logs user input upc
            message = f"User Input UPC: {upc}"
            logging.debug(message)
            
            print()
            if verifySkuAndUpc(sku, upc):
                print(Back.GREEN + Fore.WHITE + Style.BRIGHT + '✔ ITEM VERIFIED ' + Style.RESET_ALL)
                message = 'Verified!'
                logging.info(message)
                verified = True
            elif upc !='0':
                print(Back.RED + Fore.WHITE + Style.BRIGHT + '✗ WRONG ITEM SCANNED ' + Style.RESET_ALL)
                message = 'Wrong Item Scanned'
                logging.info(message)
                # verified = False
            print()

        if upc == '0':
            print()
            continue

        if verified:
            adjustment = int(quantity) * -1
            location = locationList[int(selected_loc)-1]['loc']
            updateInventory(location, sku, adjustment)
            recordShipment(sku, int(quantity))
            inventoryReportGenerator()
            print()
            print(Fore.GREEN + '✔'  + Style.RESET_ALL + ' Inventory table updated')
            print(Fore.GREEN + '✔'  + Style.RESET_ALL + ' Shipment recorded')
            print(Fore.GREEN + '✔'  + Style.RESET_ALL + ' Inventory report refreashed')
            print()
            print()

    print(Style.RESET_ALL)


def validateFile():
    skuList1 = []
    ErrorList = []
    error = {}
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    sheet = wb['SKU-UPC']
    inventorySheet = wb['Inventory']

    # while loop returns a list of SKU from the SKU-UPC workbook
    r = 2
    skuCol1 = 1

    while sheet.cell(row = r, column = skuCol1).value != None:
        sku1 = sheet.cell(row = r, column = skuCol1).value
        skuList1.append(sku1)
        r += 1

    # handles all error parsing

        # range from row 2 to the final row depending on if column A or B has more rows
    for num_row in range(2, max(len(inventorySheet["B"]), len(inventorySheet["A"]))):
        # defines 3 variables using columns 1, 2, 3
        loc = inventorySheet.cell(row = num_row, column = 1).value
        sku = inventorySheet.cell(row = num_row, column = 2).value
        quantity = inventorySheet.cell(row = num_row, column = 3).value

        # finds the location associated with each row
        while loc == None:
            num_row -= 1
            loc = inventorySheet.cell(row = num_row, column = 1).value
            
        # finds missing SKU, i.e. if SKU in inventory list but not in master sheet
        if sku != None and sku not in skuList1:
            error['type'] = 'Invalid SKU'
            error['loc'] = loc
            error['SKU'] = sku
            ErrorList.append(error)
            error = {}

        # finds if there is a missing SKU where there is a quantity
        if sku == None and quantity != None:
            error['type'] = 'Missing SKU'
            error['loc'] = loc
            error['SKU'] = None
            ErrorList.append(error)
            error = {}

        # finds if there is a missing quantity where there is a SKU
        if sku != None and quantity == None:
            error['type'] = 'Missing quantity'
            error['loc'] = loc
            error['SKU'] = sku
            ErrorList.append(error)
            error = {}
        
        # finds if the quantity inputed is not a number
        if quantity != None:
            try:
                int(quantity)
            except ValueError:
                error['type'] = 'Quantity not a number'
                error['loc'] = loc
                error['SKU'] = sku
                ErrorList.append(error)
                error = {}
        
        #TODO check if all location except 'Office' is in the correct format -> ^[A-Z]-[0-9]-[0-9] <- use regular expression

    return ErrorList


def inventoryReportGenerator():
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    inventorySheet = wb['Inventory']

    r = 2
    locCol = 1
    skuCol = 2
    qtyCol = 3

    inventoryList = {}

    while inventorySheet.cell(row = r, column = locCol).value != 'END':
        sku = inventorySheet.cell(row = r, column = skuCol).value
        qty = inventorySheet.cell(row = r, column = qtyCol).value

        if sku != None and qty != None:
            if sku in inventoryList:
                inventoryList[sku] += qty
                
            else:
                inventoryList[sku] = qty

        r += 1

    skuSheet = wb['SKU-UPC']
    r = 2
    skuCol = 1

    #put 0 in the dict if the sku is not in the inventory
    while skuSheet.cell(row = r, column = skuCol).value != None:
        sku = skuSheet.cell(row = r, column = skuCol).value
        if sku not in inventoryList:
            inventoryList[sku] = 0
            message = f" {sku} not in inventory"
        r += 1

    inventoryList = dict(sorted(inventoryList.items()))
    work_book = Workbook()
    work_sheet = work_book.active

    work_sheet.column_dimensions['A'].width = 20
    work_sheet.freeze_panes = 'A2'
    work_sheet.cell(row = 1, column = 1).value = 'SKU'
    work_sheet.cell(row = 1, column = 2).value = 'Qty'
    

    r = 2
    for key in inventoryList:
        work_sheet.cell(row = r, column = 1).value = key
        work_sheet.cell(row = r, column = 2).value = inventoryList[key]
        r += 1

    today = datetime.date.today()
    currentDate = today.strftime("%Y-%m-%d")
    fileName = './Reports/Inventory report ' + currentDate + '.xlsx'
    work_book.save(fileName)


# def shipmentReportGenerator():
#     with open('./csv/amazon.csv', mode="r") as csv_file:
#         reader = csv.reader(csv_file)
#         current_time = datetime.datetime.now()

#         today = numberToDateStr(current_time.month, 11)

#         ordersOfToday = []
#         for item in reader:
#             dateTimeStr = item[0]
#             dateStr = dateTimeStr[:dateTimeStr.find(',')]
#             if dateStr == today:
#                 ordersOfToday.append(item)
        
#         fbmOrders = []
#         for order in ordersOfToday:
#             if 'FBM' in order[4]:
#                 fbmOrders.append(order)
        
#         for order in fbmOrders:
#             orderObj = {}
#             orderObj['orderNumber'] = order[3]
#             orderObj['sku'] = order[4]
            

def numberToDateStr(month, day):
    if month == 1:
        date = 'January'
    elif month == 2:
        date = 'February'
    elif month == 3:
        date = 'March'
    elif month == 4:
        date = 'April'
    elif month == 5:
        date = 'May'
    elif month == 6:
        date = 'June'
    elif month == 7:
        date = 'July'
    elif month == 8:
        date = 'Eight'
    elif month == 9:
        date = 'September'
    elif month == 10:
        date = 'October'
    elif month == 11:
        date = 'November'
    elif month == 12:
        date = 'December'
    
    return date + ' ' + str(day)


def searchInventory(skuInput):
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    sheet = wb['Inventory']

    r = 2
    locCol = 1
    skuCol = 2
    qtyCol = 3

    locationList = []

    while sheet.cell(row = r, column = locCol).value != 'END':
        loc = sheet.cell(row = r, column = locCol).value
        sku = sheet.cell(row = r, column = skuCol).value
        qty = sheet.cell(row = r, column = qtyCol).value

        #get the location value
        if sku != None and sku == skuInput:
            i = 1
            while loc == None:
                loc = sheet.cell(row = r - i, column = locCol).value
                i += 1

            location = {
                'sku': sku,
                'qty': qty,
                'loc': loc,
            }

            locationList.append(location)

        r += 1

    if len(locationList) > 1:
        for i in range(len(locationList) - 1):
            for j in range (i + 1, len(locationList)):
                if comapreLoc(locationList[i]['loc'], locationList[j]['loc']):
                    temp = locationList[i]
                    locationList[i] = locationList[j]
                    locationList[j] = temp

    return locationList


def parseSKU(itemNumber):
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    sheet = wb['SKU Alies']

    r = 2
    aliesCol = 1
    skuCol = 2

    while sheet.cell(row = r, column = aliesCol).value != None:
        if str(itemNumber) == str(sheet.cell(row = r, column = aliesCol).value):
            return sheet.cell(row = r, column = skuCol).value
        r += 1
    return itemNumber


#find the closer lcoation
def comapreLoc(loc1, loc2):
    loc1Arr = loc1.split('-')
    loc2Arr = loc2.split('-')

    if loc1 == 'Office': return False
    if loc2 == 'Office': return True

    if loc1Arr[2] > loc2Arr[2]: return True
    if loc1Arr[2] < loc2Arr[2]: return False
    if loc1Arr[1] > loc2Arr[1]: return True
    if loc1Arr[1] < loc2Arr[1]: return False
    if loc1Arr[0] > loc2Arr[0]: return True
    if loc1Arr[0] < loc2Arr[0]: return False
    return False


#check the the workbook 'SKU-UPC' in the excel file to verify if this UPC is correct to the SKU
def verifySkuAndUpc(skuInput, upc_input):
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    sheet = wb['SKU-UPC']

    r = 2
    skuCol = 1
    upcCol = 2

    while sheet.cell(row = r, column = skuCol).value != None:
        sku = sheet.cell(row = r, column = skuCol).value

        if skuInput == sku:
            upc = sheet.cell(row = r, column = upcCol).value
            if upc_input == str(upc):
                return True
            else:
                return False
        
        r += 1


def updateInventory(location, itemSku, adjustment):
    wb = load_workbook('./Excel/Inventory by location.xlsx')
    sheet = wb['Inventory']

    r = 2
    locCol = 1
    skuCol = 2
    qtyCol = 3

    while sheet.cell(row = r, column = locCol).value != 'END':
        loc = sheet.cell(row = r, column = locCol).value
        if loc == location:
            sku = sheet.cell(row = r, column = skuCol).value

            while itemSku != sku:
                r += 1
                sku = sheet.cell(row = r, column = skuCol).value

            qty = sheet.cell(row = r, column = qtyCol)
            message = f'Initial quantity: {qty.value}'
            logging.info(message)
            qty.value = qty.value + adjustment
            message = f"Updated quantity: {qty.value}"
            logging.info(message)
            
            #if quantity is 0, then the entire row's values will be deleted
            if qty.value == 0:
                sheet.cell(row = r, column = skuCol).value = None
                qty.value = None
                message = f'{loc}:{sku} -> Quantity is 0 so row {r} deleted'
                logging.info(message)
                
            wb.save('./Excel/Inventory by location.xlsx')
            break
        r += 1


def recordShipment(sku, quantity):
    # print('triggered')
    wb = load_workbook('./Excel/Daily shipment (version 1).xlsx')
    sheet = wb['SHP Items']

    r = 1
    dateCol = 1
    skuCol = 2
    fbmCol = 4

    while sheet.cell(row = r, column = dateCol).value != 'END':
        r += 1

    lastDateRow = r

    lastDateRow -= 1
    while sheet.cell(row = lastDateRow, column = dateCol).value == None:
        lastDateRow -= 1

    lastDate = sheet.cell(row = lastDateRow, column = dateCol).value.date()

    today = datetime.date.today()

    if today == lastDate:
        existingSkuRow = None

        for i in range(lastDateRow, r):
            if sheet.cell(row = i, column = skuCol).value == sku:
                existingSkuRow = i
                sheet.cell(row = existingSkuRow, column = fbmCol).value += quantity
                break

        if existingSkuRow == None:
            sheet.insert_rows(r, 1)
            sheet.cell(row = r, column = skuCol).value = sku
            sheet.cell(row = r, column = fbmCol).value = quantity
            message = f'Added {sku} with {quantity} quantity to row {r}'
            logging.info(message)

    else:
        sheet.insert_rows(r, 2)
        r += 1
        sheet.cell(row = r, column = dateCol).value = today
        sheet.cell(row = r, column = skuCol).value = sku
        sheet.cell(row = r, column = fbmCol).value = quantity
        message = f'Added {sku} with {quantity} quantity to new row {r}'
        logging.info(message)

    wb.save('./Excel/Daily shipment (version 1).xlsx')

if __name__ == '__main__':
    main()